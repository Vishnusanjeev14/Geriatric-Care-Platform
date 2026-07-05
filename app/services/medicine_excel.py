from pathlib import Path

from openpyxl import load_workbook


CATALOGUE_PATH = Path(__file__).resolve().parents[2] / "data" / "medicine_master.xlsx"


def load_master_medicines(path=CATALOGUE_PATH):
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook["Medicine Master"]
    headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    rows = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        rows.append(dict(zip(headers, row)))
    workbook.close()
    return rows
