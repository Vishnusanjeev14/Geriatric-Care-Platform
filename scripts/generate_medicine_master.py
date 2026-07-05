from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo


OUTPUT = Path(__file__).resolve().parents[1] / "data" / "medicine_master.xlsx"


BASE_MEDICINES = [
    ("Paracetamol", "Acetaminophen", "Pain Relief", "Tablet", False),
    ("Aspirin", "Acetylsalicylic Acid", "Pain Relief", "Tablet", False),
    ("Ibuprofen", "Ibuprofen", "Pain Relief", "Tablet", False),
    ("Amoxicillin", "Amoxicillin", "Antibiotics", "Capsule", True),
    ("Azithromycin", "Azithromycin", "Antibiotics", "Tablet", True),
    ("Metformin", "Metformin", "Diabetes Care", "Tablet", True),
    ("Amlodipine", "Amlodipine", "Heart Health", "Tablet", True),
    ("Atorvastatin", "Atorvastatin", "Heart Health", "Tablet", True),
    ("Losartan", "Losartan Potassium", "Heart Health", "Tablet", True),
    ("Omeprazole", "Omeprazole", "Digestive Health", "Capsule", False),
    ("Pantoprazole", "Pantoprazole", "Digestive Health", "Tablet", False),
    ("Cetirizine", "Cetirizine", "Allergy", "Tablet", False),
    ("Loratadine", "Loratadine", "Allergy", "Tablet", False),
    ("Salbutamol", "Albuterol", "Respiratory", "Inhaler", True),
    ("Montelukast", "Montelukast", "Respiratory", "Tablet", True),
    ("Levothyroxine", "Levothyroxine", "Endocrine", "Tablet", True),
    ("Vitamin D3", "Cholecalciferol", "Wellness", "Capsule", False),
    ("Calcium Citrate", "Calcium Citrate", "Wellness", "Tablet", False),
    ("Ferrous Sulfate", "Ferrous Sulfate", "Wellness", "Tablet", False),
    ("Insulin Glargine", "Insulin Glargine", "Diabetes Care", "Pen", True),
    ("Dextromethorphan", "Dextromethorphan", "Cough & Cold", "Syrup", False),
]

STRENGTHS = ["100 mg", "250 mg", "500 mg", "650 mg", "5 mg", "10 mg", "20 mg", "40 mg", "1000 IU", "2 mg/ml"]
BRANDS = ["CareWell", "NovaHealth", "MediCore", "AstraLife", "WellSpring", "CuraPlus", "NorthStar", "Vitalis"]
MANUFACTURERS = ["Global Pharma", "Apex Laboratories", "Zenith Health", "Evergreen Medical", "Prime Therapeutics", "BluePeak Pharma"]
PACK_BY_FORM = {
    "Tablet": "10 tablets / strip",
    "Capsule": "10 capsules / strip",
    "Syrup": "100 ml bottle",
    "Inhaler": "200 dose inhaler",
    "Pen": "3 ml prefilled pen",
}

BASE_PRICE_BY_MEDICINE = {
    "Paracetamol": 18,
    "Aspirin": 26,
    "Ibuprofen": 24,
    "Amoxicillin": 115,
    "Azithromycin": 92,
    "Metformin": 28,
    "Amlodipine": 36,
    "Atorvastatin": 82,
    "Losartan": 74,
    "Omeprazole": 58,
    "Pantoprazole": 72,
    "Cetirizine": 22,
    "Loratadine": 48,
    "Salbutamol": 185,
    "Montelukast": 145,
    "Levothyroxine": 110,
    "Vitamin D3": 130,
    "Calcium Citrate": 165,
    "Ferrous Sulfate": 42,
    "Insulin Glargine": 760,
    "Dextromethorphan": 96,
}

FORM_PRICE_MULTIPLIER = {
    "Tablet": 1,
    "Capsule": 1.08,
    "Syrup": 1.6,
    "Inhaler": 1.35,
    "Pen": 1.85,
}

STRENGTH_PRICE_MULTIPLIER = {
    "100 mg": 0.72,
    "250 mg": 0.85,
    "500 mg": 1,
    "650 mg": 1.12,
    "5 mg": 0.88,
    "10 mg": 1,
    "20 mg": 1.15,
    "40 mg": 1.35,
    "1000 IU": 1.05,
    "2 mg/ml": 1.2,
}


def realistic_price(name, form, strength, index):
    base = BASE_PRICE_BY_MEDICINE[name]
    form_multiplier = FORM_PRICE_MULTIPLIER[form]
    strength_multiplier = STRENGTH_PRICE_MULTIPLIER[strength]
    brand_variation = 1 + ((index % len(BRANDS)) - 3) * 0.035
    return round(max(6, base * form_multiplier * strength_multiplier * brand_variation), 2)


def build_rows(target_count=1200):
    rows = []
    for index in range(target_count):
        name, generic, category, form, rx = BASE_MEDICINES[index % len(BASE_MEDICINES)]
        strength = STRENGTHS[(index // len(BASE_MEDICINES)) % len(STRENGTHS)]
        brand = f"{BRANDS[index % len(BRANDS)]} {name}"
        manufacturer = MANUFACTURERS[index % len(MANUFACTURERS)]
        price = realistic_price(name, form, strength, index)
        stock = 30 + ((index * 17) % 970)
        rows.append(
            [
                f"{name} {strength}",
                generic,
                brand,
                strength,
                form,
                price,
                PACK_BY_FORM[form],
                manufacturer,
                category,
                "Yes" if rx else "No",
                stock,
                25 if rx else 40,
                f"89010{index + 100000:07d}",
                "2027-12",
            ]
        )
    return rows


def main():
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Medicine Master"
    headers = [
        "Medicine Name",
        "Generic Name",
        "Brand",
        "Strength",
        "Form",
        "Price",
        "Pack Size",
        "Manufacturer",
        "Category",
        "Prescription Required",
        "Available Stock",
        "Minimum Stock",
        "Barcode",
        "Expiry Placeholder",
    ]
    sheet.append(headers)
    for row in build_rows():
        sheet.append(row)

    header_fill = PatternFill("solid", fgColor="0F766E")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")

    widths = [28, 24, 28, 12, 14, 10, 18, 24, 20, 22, 16, 14, 18, 18]
    for column, width in enumerate(widths, start=1):
        sheet.column_dimensions[chr(64 + column)].width = width

    table = Table(displayName="MedicineMasterTable", ref=f"A1:N{sheet.max_row}")
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium4", showRowStripes=True)
    sheet.add_table(table)
    sheet.freeze_panes = "A2"
    workbook.save(OUTPUT)
    print(OUTPUT)


if __name__ == "__main__":
    main()
